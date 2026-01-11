"""
Attachment handler for downloading and processing email attachments.
Supports various file types including documents, images, and PDFs.
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from email import policy
from email.parser import BytesParser
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
from src.utils import sanitize_filename, ensure_directory, get_file_size_readable


class AttachmentHandler:
    """
    Handler for email attachments - download, save, and extract content.
    """
    
    def __init__(self, base_save_path: str = "SavedDocuments/attachments"):
        """
        Initialize attachment handler.
        
        Args:
            base_save_path: Base directory for saving attachments
        """
        self.logger = logging.getLogger(__name__)
        self.base_save_path = base_save_path
        ensure_directory(base_save_path)
    
    def process_attachments(self, msg, email_id: str) -> List[Dict[str, Any]]:
        """
        Process all attachments in an email.
        
        Args:
            msg: Email message object
            email_id: Unique identifier for the email
            
        Returns:
            List of attachment information dictionaries
        """
        attachments_info = []
        
        if not msg.is_multipart():
            return attachments_info
        
        # Create directory for this email's attachments
        email_attachment_dir = os.path.join(self.base_save_path, email_id)
        ensure_directory(email_attachment_dir)
        
        for part in msg.walk():
            content_disposition = str(part.get("Content-Disposition", ""))
            
            if "attachment" in content_disposition:
                attachment_info = self._save_attachment(part, email_attachment_dir)
                if attachment_info:
                    attachments_info.append(attachment_info)
        
        self.logger.info(f"Processed {len(attachments_info)} attachments for email {email_id}")
        return attachments_info
    
    def _save_attachment(self, part, save_dir: str) -> Optional[Dict[str, Any]]:
        """
        Save individual attachment.
        
        Args:
            part: Email part containing attachment
            save_dir: Directory to save the attachment
            
        Returns:
            Attachment information dictionary or None if failed
        """
        try:
            filename = part.get_filename()
            
            if not filename:
                self.logger.warning("Attachment without filename, skipping")
                return None
            
            # Decode filename if needed
            if isinstance(filename, bytes):
                filename = filename.decode()
            
            # Sanitize filename
            filename = sanitize_filename(filename)
            filepath = os.path.join(save_dir, filename)
            
            # Get attachment data
            attachment_data = part.get_payload(decode=True)
            
            if not attachment_data:
                self.logger.warning(f"No data for attachment: {filename}")
                return None
            
            # Save file
            with open(filepath, 'wb') as f:
                f.write(attachment_data)
            
            file_size = len(attachment_data)
            content_type = part.get_content_type()
            
            attachment_info = {
                "filename": filename,
                "filepath": filepath,
                "size": file_size,
                "size_readable": get_file_size_readable(file_size),
                "type": content_type
            }
            
            # Try to extract text content
            text_content = self._extract_text_content(filepath, content_type)
            if text_content:
                attachment_info["text_content"] = text_content
            
            self.logger.info(f"Saved attachment: {filename} ({attachment_info['size_readable']})")
            return attachment_info
            
        except Exception as e:
            self.logger.error(f"Error saving attachment: {e}")
            return None
    
    def _extract_text_content(self, filepath: str, content_type: str) -> Optional[str]:
        """
        Extract text content from attachments.
        
        Args:
            filepath: Path to the attachment file
            content_type: MIME type of the file
            
        Returns:
            Extracted text or None
        """
        try:
            # PDF files
            if content_type == 'application/pdf' or filepath.endswith('.pdf'):
                return self._extract_from_pdf(filepath)
            
            # Word documents
            elif content_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' \
                or filepath.endswith('.docx'):
                return self._extract_from_docx(filepath)
            
            # Excel files
            elif content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' \
                or filepath.endswith('.xlsx'):
                return self._extract_from_xlsx(filepath)
            
            # Plain text files
            elif content_type.startswith('text/') or filepath.endswith('.txt'):
                return self._extract_from_text(filepath)
            
            # Images (basic info only)
            elif content_type.startswith('image/'):
                return self._get_image_info(filepath)
            
            return None
            
        except Exception as e:
            self.logger.warning(f"Failed to extract content from {filepath}: {e}")
            return None
    
    def _extract_from_pdf(self, filepath: str) -> str:
        """Extract text from PDF file."""
        try:
            text_content = []
            with open(filepath, 'rb') as f:
                pdf_reader = PdfReader(f)
                for page in pdf_reader.pages:
                    text_content.append(page.extract_text())
            return '\n'.join(text_content)
        except Exception as e:
            self.logger.warning(f"Error extracting PDF: {e}")
            return ""
    
    def _extract_from_docx(self, filepath: str) -> str:
        """Extract text from Word document."""
        try:
            doc = Document(filepath)
            return '\n'.join([para.text for para in doc.paragraphs])
        except Exception as e:
            self.logger.warning(f"Error extracting DOCX: {e}")
            return ""
    
    def _extract_from_xlsx(self, filepath: str) -> str:
        """Extract text from Excel file."""
        try:
            workbook = load_workbook(filepath, read_only=True)
            text_content = []
            
            for sheet in workbook.worksheets:
                text_content.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                    text_content.append(row_text)
                text_content.append('')
            
            return '\n'.join(text_content)
        except Exception as e:
            self.logger.warning(f"Error extracting XLSX: {e}")
            return ""
    
    def _extract_from_text(self, filepath: str) -> str:
        """Extract content from text file."""
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            self.logger.warning(f"Error reading text file: {e}")
            return ""
    
    def _get_image_info(self, filepath: str) -> str:
        """Get basic information about image file."""
        try:
            with Image.open(filepath) as img:
                return f"Image: {img.format}, Size: {img.size}, Mode: {img.mode}"
        except Exception as e:
            self.logger.warning(f"Error reading image: {e}")
            return "Image file"
    
    def get_attachment_summary(self, attachments: List[Dict[str, Any]]) -> str:
        """
        Generate a summary of attachments.
        
        Args:
            attachments: List of attachment information dictionaries
            
        Returns:
            Formatted summary string
        """
        if not attachments:
            return "No attachments"
        
        total_size = sum(att['size'] for att in attachments)
        summary = f"{len(attachments)} attachment(s), Total size: {get_file_size_readable(total_size)}\n"
        
        for i, att in enumerate(attachments, 1):
            summary += f"{i}. {att['filename']} ({att['size_readable']}) - {att['type']}\n"
        
        return summary
